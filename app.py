import streamlit as st
import pandas as pd
import os
import random
from fpdf import FPDF

# Load the data
file_path = 'C:/Mira/LEMME_Chat_Translated_Manual_DE.xlsx'
if not os.path.exists(file_path):
    st.error("Die Datei wurde nicht gefunden. Bitte stellen Sie sicher, dass sich die Datei unter 'C:/Mira/LEMME_Chat_Translated_Manual_DE.xlsx' befindet.")
    st.stop()

try:
    # Ensure openpyxl is available
    import openpyxl
    data = pd.read_excel(file_path, sheet_name='Sheet1')
except ImportError:
    st.error("Fehlende Abhängigkeit: 'openpyxl'. Installieren Sie es mit 'pip install openpyxl'.")
    st.stop()
except Exception as e:
    st.error(f"Fehler beim Laden der Datei: {e}")
    st.stop()

# Clean the data to remove unnecessary rows (e.g., headers within the data)
data = data.dropna(subset=['Frühstück', 'Mittag', 'Abend'])
if data.empty:
    st.error("Die Daten sind leer oder ungültig. Bitte überprüfen Sie die Quelle.")
    st.stop()

# Normalize data to avoid filtering issues
data['Frühstück'] = data['Frühstück'].astype(str).str.strip().str.lower()
data['Mittag'] = data['Mittag'].astype(str).str.strip()
data['Abend'] = data['Abend'].astype(str).str.strip()

# Memory for previously used meals
used_meals = set()

# Function to generate PDF
def generate_pdf(plan, title):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=title, ln=True, align='C')
    pdf.ln(10)

    for day, meals in plan.items():
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(200, 10, txt=f"{day}", ln=True)
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"- Frühstück: {meals['Frühstück']}", ln=True)
        pdf.cell(200, 10, txt=f"- Mittag: {meals['Mittag']}", ln=True)
        pdf.cell(200, 10, txt=f"- Abend: {meals['Abend']}", ln=True)
        pdf.ln(5)

    return pdf

# Streamlit App
def main():
    st.title("Mahlzeit-Empfehlungen mit Wochenplan und Einkaufsliste")

    # Wochenplan erstellen mit Tabs
    st.write("## Wochenplan erstellen")
    days = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']
    weekly_plan = {}

    tabs = st.tabs(days)

    for day, tab in zip(days, tabs):
        with tab:
            st.write(f"### {day}")

            # Auswahl der Startmahlzeit
            starting_meal = st.radio(f"Mit welcher Mahlzeit möchtest du für {day} beginnen?", ("Frühstück", "Mittag", "Abend"), key=f"starting_meal_{day}")

            # Frühstücksauswahl
            selected_breakfast = None
            selected_lunch = None
            selected_dinner = None

            if starting_meal == "Frühstück":
                selected_breakfast = st.selectbox(f"Wähle ein Frühstück für {day}", data['Frühstück'].unique(), key=f"breakfast_{day}")
                matching_lunch = data[data['Frühstück'] == selected_breakfast]['Mittag'].unique()
                selected_lunch = st.selectbox(f"Mittagessen basierend auf deiner Frühstücksauswahl ({day}):", matching_lunch, key=f"lunch_{day}")
                matching_dinner = data[(data['Frühstück'] == selected_breakfast) & (data['Mittag'] == selected_lunch)]['Abend'].unique()
                selected_dinner = st.selectbox(f"Abendessen basierend auf deiner Frühstücks- und Mittagsauswahl ({day}):", matching_dinner, key=f"dinner_{day}")
            elif starting_meal == "Mittag":
                selected_lunch = st.selectbox(f"Wähle ein Mittagessen für {day}", data['Mittag'].unique(), key=f"lunch_{day}")
                matching_breakfast = data[data['Mittag'] == selected_lunch]['Frühstück'].unique()
                selected_breakfast = st.selectbox(f"Frühstück basierend auf deiner Mittagsauswahl ({day}):", matching_breakfast, key=f"breakfast_{day}")
                matching_dinner = data[(data['Mittag'] == selected_lunch) & (data['Frühstück'] == selected_breakfast)]['Abend'].unique()
                selected_dinner = st.selectbox(f"Abendessen basierend auf deiner Frühstücks- und Mittagsauswahl ({day}):", matching_dinner, key=f"dinner_{day}")
            elif starting_meal == "Abend":
                selected_dinner = st.selectbox(f"Wähle ein Abendessen für {day}", data['Abend'].unique(), key=f"dinner_{day}")
                matching_lunch = data[data['Abend'] == selected_dinner]['Mittag'].unique()
                selected_lunch = st.selectbox(f"Mittagessen basierend auf deiner Abendauswahl ({day}):", matching_lunch, key=f"lunch_{day}")
                matching_breakfast = data[(data['Abend'] == selected_dinner) & (data['Mittag'] == selected_lunch)]['Frühstück'].unique()
                selected_breakfast = st.selectbox(f"Frühstück basierend auf deiner Mittag- und Abendauswahl ({day}):", matching_breakfast, key=f"breakfast_{day}")

            weekly_plan[day] = {
                'Frühstück': selected_breakfast,
                'Mittag': selected_lunch,
                'Abend': selected_dinner
            }

    # Warnung, wenn eine Mahlzeit mehr als 2-mal ausgewählt wurde
    all_selected_meals = [meal for day_meals in weekly_plan.values() for meal in day_meals.values() if meal]
    meal_counts = pd.Series(all_selected_meals).value_counts()
    repeated_meals = meal_counts[meal_counts > 2]
    if not repeated_meals.empty:
        st.warning("Folgende Mahlzeiten wurden mehr als 2-mal in der Woche ausgewählt und sollten reduziert werden:")
        for meal, count in repeated_meals.items():
            st.write(f"- {meal}: {count}x")

    # Automatischen Wochenplan generieren
    if st.button("Automatischen Wochenplan erstellen"):
        auto_plan = {}
        available_meals = data.copy()

        for day in days:
            day_plan = {}
            for meal_type in ['Frühstück', 'Mittag', 'Abend']:
                remaining_meals = available_meals[meal_type][~available_meals[meal_type].isin(used_meals)].unique()
                if len(remaining_meals) == 0:
                    st.error(f"Nicht genügend verfügbare {meal_type} Optionen für {day}.")
                    return

                selected_meal = random.choice(remaining_meals)
                day_plan[meal_type] = selected_meal

                # Mahlzeit als verwendet markieren
                used_meals.add(selected_meal)

                # Sicherstellen, dass maximal 2 Tage hintereinander gleiche Mahlzeiten gewählt werden
                if len(used_meals) > 7:
                    used_meals.pop()

            auto_plan[day] = day_plan

        st.write("## Automatisch erstellter Wochenplan")
        for day, meals in auto_plan.items():
            st.write(f"### {day}")
            st.write(f"- **Frühstück:** {meals['Frühstück']}")
            st.write(f"- **Mittag:** {meals['Mittag']}")
            st.write(f"- **Abend:** {meals['Abend']}")

        # PDF generieren und herunterladen
        pdf = generate_pdf(auto_plan, "Automatisch erstellter Wochenplan")
        pdf_output = "wochenplan.pdf"
        pdf.output(pdf_output)
        with open(pdf_output, "rb") as pdf_file:
            st.download_button("Wochenplan als PDF herunterladen", data=pdf_file, file_name="Wochenplan.pdf", mime="application/pdf")

    # Automatischen Monatsplan generieren
    if st.button("Automatischen Monatsplan erstellen"):
        days_in_month = 30
        auto_month_plan = {}
        available_meals = data.copy()

        for day in range(1, days_in_month + 1):
            day_plan = {}
            for meal_type in ['Frühstück', 'Mittag', 'Abend']:
                remaining_meals = available_meals[meal_type][~available_meals[meal_type].isin(used_meals)].unique()
                if len(remaining_meals) == 0:
                    st.error(f"Nicht genügend verfügbare {meal_type} Optionen für Tag {day}.")
                    return

                selected_meal = random.choice(remaining_meals)
                day_plan[meal_type] = selected_meal

                # Mahlzeit als verwendet markieren
                used_meals.add(selected_meal)

                # Sicherstellen, dass maximal 2 Tage hintereinander gleiche Mahlzeiten gewählt werden
                if len(used_meals) > 7:
                    used_meals.pop()

            auto_month_plan[f"Tag {day}"] = day_plan

        st.write("## Automatisch erstellter Monatsplan")
        for day, meals in auto_month_plan.items():
            st.write(f"### {day}")
            st.write(f"- **Frühstück:** {meals['Frühstück']}")
            st.write(f"- **Mittag:** {meals['Mittag']}")
            st.write(f"- **Abend:** {meals['Abend']}")

        # PDF generieren und herunterladen
        pdf = generate_pdf(auto_month_plan, "Automatisch erstellter Monatsplan")
        pdf_output = "monatsplan.pdf"
        pdf.output(pdf_output)
        with open(pdf_output, "rb") as pdf_file:
            st.download_button("Monatsplan als PDF herunterladen", data=pdf_file, file_name="Monatsplan.pdf", mime="application/pdf")

    # Wochenplan anzeigen
    if st.button("Wochenplan anzeigen"):
        st.write("## Dein Wochenplan (Druckversion)")
        for day, meals in weekly_plan.items():
            st.write(f"### {day}")
            st.write(f"- **Frühstück:** {meals['Frühstück']}")
            st.write(f"- **Mittag:** {meals['Mittag']}")
            st.write(f"- **Abend:** {meals['Abend']}")

if __name__ == "__main__":
    main()
